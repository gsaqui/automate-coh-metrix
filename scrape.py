#!/usr/bin/python

import sys
import os.path
from os import listdir
from os.path import isfile, join

import requests

from lxml import html
from openpyxl import Workbook
from openpyxl import load_workbook


# Console text colours
class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'


# Posts the data to Coh Metrix website
def get_data(writingSample, code, genre):
    s = requests.Session()
    starter = s.get('http://141.225.42.101/cohmetrix3/login.aspx', timeout=120)
    tree = html.fromstring(starter.text)

    cookies = {
        'ASP.NET_SessionId': starter.cookies['ASP.NET_SessionId']
    }

    headers = {
        'Host': '141.225.42.101',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.10; rv:41.0) Gecko/20100101 Firefox/41.0',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Referer': 'http://141.225.42.101/cohmetrix3/login.aspx',
        'Connection': 'keep-alive',
        'Pragma': 'no-cache',
        'Cache-Control': 'no-cache',
    }
    viewstate = tree.xpath('//input[@id="__VIEWSTATE"]/@value')[0]
    validation = tree.xpath('//input[@id="__VIEWSTATEGENERATOR"]/@value')[0]
    view = tree.xpath('//input[@id="__EVENTVALIDATION"]/@value')[0]

    data = {
        '__VIEWSTATE': viewstate,
        '__VIEWSTATEGENERATOR': validation,
        '__EVENTVALIDATION': view,
        'TextBox1': 'glenn.saqui@gmail.com',
        'TextBox2': "bebZjcis",
        'Button1': 'Login',
    }
    l = s.post('http://141.225.42.101/cohmetrix3/login.aspx', headers=headers, data=data, cookies=cookies, timeout=120)

    headers = {
        'Host': '141.225.42.101',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.10; rv:41.0) Gecko/20100101 Firefox/41.0',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Referer': 'http://141.225.42.101/cohmetrix3/input.aspx',
        'Connection': 'keep-alive',
        'Pragma': 'no-cache',
        'Cache-Control': 'no-cache',
    }
    viewstate = "/wEPDwUJNzU4MDYzOTUyZGQKNmk5nIq1XYBgXUM1EZ1Hs0bJWQ=="
    validation = '/wEWCgLrrJCoDQK506zcAwLy07LaDALjz837AwLBx6n3AwLptdyVCAL0geaDCwLCoveuCALpouS8DQLCi9reA0rdDT2rbbn4VrlhjGW0OKnV7hBW'

    sampleData = {
        '__VIEWSTATE': viewstate,
        '__VIEWSTATEGENERATOR': 'AD710423',
        '__EVENTVALIDATION': validation,
        'tbTitle': 'test',
        'ddlGenre': genre,
        'tbSource': 'source',
        'tbUserCode': code,
        'ddlLSASpace': 'CollegeLevel',
        'tbInput': writingSample,
        'btnSubmit': 'Submit',
    }

    t = s.post('http://141.225.42.101/cohmetrix3/input.aspx', headers=headers, cookies=cookies, data=sampleData, timeout=120)
    return t.text


# parses the data from the html and sticks it into the excel spreadsheet
def parse_data(text, writingSampleId, outputFileName):
    tree = html.fromstring(text)
    rows = tree.xpath('//tr')
    workbook = get_result_file(outputFileName)

    worksheet = workbook.active
    worksheet.title = "Coh Results"

    add_excel_header(rows, worksheet)

    # Read all the rows and check to make sure we haven't put this value in before
    add_parsed_results_to_spreadsheet(rows, worksheet, writingSampleId)

    workbook.save(outputFileName)


# takes the rows that are parsed from the webiste and insert them into the excel spreadsheet
def add_parsed_results_to_spreadsheet(rows, worksheet, writingSampleId):
    newRowNumber = len(worksheet.rows) + 1
    columnNumber = 1
    worksheet.cell(row=newRowNumber, column=columnNumber).value = writingSampleId
    for row in rows:
        cols = row.getchildren()
        if len(cols) == 5:
            columnNumber += 1
            worksheet.cell(row=newRowNumber, column=columnNumber).value = cols[3].text


# Adds the excel column header
def add_excel_header(rows, worksheet):
    columnNumber = 1
    if worksheet['A1'].value != 'ID':
        worksheet.cell(row=1, column=columnNumber).value = 'ID'

        for row in rows:
            cols = row.getchildren()
            if len(cols) == 5:
                columnNumber += 1
                worksheet.cell(row=1, column=columnNumber).value = cols[1].text


# Returns a workbook either by creating a new one or by reading the previous results
def get_result_file(fileName):
    if os.path.isfile(fileName):
        return load_workbook(fileName)

    return Workbook()


# Checks all the files in the writing-samples directory to check to see if there are any new ones that need to be run
def get_files_to_send_results_for(sampleLocation, outputFileName):
    filesToBeUploaded = []

    if os.path.isdir(sampleLocation):
        onlyfiles = [f for f in listdir(sampleLocation) if isfile(join(sampleLocation, f))]

        workbook = get_result_file(outputFileName)
        worksheet = workbook.active

        for file in onlyfiles:
            sampleId = file.split('.')[0]

            shouldSampleBeAdded = True
            for row in worksheet.rows:
                cellValue = str(worksheet.cell(row=row[0].row, column=1).value)
                if cellValue == sampleId:
                    shouldSampleBeAdded = False

            if shouldSampleBeAdded:
                filesToBeUploaded.append(file)

    return filesToBeUploaded


if len(sys.argv) <= 3:
    print
    print bcolors.FAIL + "Usages: python scrape.py <Genre (Science|Narrative|Informational)> <directory where all writing samples are located> <results file.xlsx>" + bcolors.ENDC
    print
    sys.exit(-1)

genre = sys.argv[1]
directoryOfWritingSamples = sys.argv[2]
outputFileName = sys.argv[3]

filesToBeUploaded = get_files_to_send_results_for(directoryOfWritingSamples, outputFileName)
print "Number of files to be uploaded: " + str(len(filesToBeUploaded))
count = 1
for filename in filesToBeUploaded:
    print str(count) + "-" + str(len(filesToBeUploaded)) + "   " + filename
    f = open(directoryOfWritingSamples + '/' + filename, 'r')
    writingSample = f.read().decode('windows-1252')
    f.close()


    writingSample = writingSample.replace(u"\u2018", "'").replace(u"\u2019", "'").replace(u"\u201c","\"").replace(u"\u201d", "\"")
    results = get_data(writingSample, filename.split('.')[0], genre)

    parse_data(results, filename.split('.')[0], outputFileName)
    count += 1

print "We are all done"
